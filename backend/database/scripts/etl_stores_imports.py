"""
etl_stores_imports.py — shared foundation for the IMPORTS and STORES loaders.

This is the imports/stores counterpart of the logistics `etl_common.py`. It
reuses that module's placeholder-aware cleaners (clean_text / clean_number /
clean_int / clean_date) so keys and values are cleaned identically everywhere,
and adds the pieces this domain needs:

  * read_report()            — open a standalone .xls / .xlsx report
  * ensure_items()           — upsert the shared `items` master
  * ensure_purchase_orders() — upsert `purchase_order`
  * load_import_map()        — {import_ref: import_id}  (FK resolution)
  * load_shipment_map()      — {(import_id, batch_no): shipment_id}

There is no `suppliers` master: the supplier name/country/city ride along as
plain TEXT columns on import_details (see schemas/imports_schemas.py).

Source workbooks are located by etl_common.data_file(), which searches this
project's own data/ folder — no machine-specific paths to adjust.

Requires:  pip install pandas openpyxl xlrd psycopg2-binary
           (xlrd is needed for the old binary .xls report exports)
"""

import json
import re
import urllib.request

import openpyxl
import pandas as pd
from psycopg2.extras import execute_values



# Reuse the exact same cleaners the logistics loaders use.
from backend.database.scripts.etl_common import (
    clean_text, clean_number, clean_int, clean_date, bulk_insert, data_file,
)

IMPORT_FILE    = data_file("imports")
STOCK_FILE     = data_file("stocks")
ISSUANCE_FILE  = data_file("issuances")
STORE_REQ_FILE = data_file("store_requisitions")

# The Import Status Sheet carries two banner rows above the real header.
IMPORT_HEADER_ROW = 2


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------

def read_report(path, header=0) -> pd.DataFrame:
    """Open a standalone report file and drop fully-empty rows.

    `header` skips banner rows some exports carry (the Import Status sheet has
    its header on the 3rd row -> header=IMPORT_HEADER_ROW).
    """
    df = pd.read_excel(path, header=header)
    df = df.dropna(how="all")
    print(f"Read '{path}': {len(df)} rows, {len(df.columns)} columns")
    return df


def read_import_rows() -> pd.DataFrame:
    """Read the Import Status Sheet and return only the *real* data rows.

    The sheet has two banner rows (title + section) above the header, plus a
    long tail of formula-only rows that carry values in a few computed columns
    but no actual import data. We keep a row only if it has an Item Name or a
    Tab Status — that isolates the genuine import rows.

    The DataFrame index (original Excel row position) is preserved and used as
    the per-import link key: this source has NO natural import id, and
    import_id is a DB-generated surrogate, so we key each import on its row
    position purely to attach the child rows (import_item / shipment_details /
    payment_history) to the right parent during loading.
    """
    df = read_report(IMPORT_FILE, header=IMPORT_HEADER_ROW)
    mask = df["Item Name"].notna() | df["Tab Status"].notna()
    real = df.loc[mask]
    print(f"  real import rows: {len(real)} (dropped {len(df) - len(real)} formula/blank rows)")
    return real


def import_ref_for(idx) -> str:
    """Stable per-row link key derived from the Excel row position. Stored in
    import_details.import_ref so the child loaders can resolve their parent
    (there is no natural import id in the source)."""
    return f"R{idx}"


# ---------------------------------------------------------------------------
# Master-table upserts (imports schema owns items / suppliers / purchase_order)
# ---------------------------------------------------------------------------
# The transaction tables carry FKs to these masters, so the masters must exist
# before the children load. Each helper is idempotent (ON CONFLICT DO NOTHING),
# so any loader can safely "ensure" the codes it needs, in any order.

ITEM_COLUMNS = [
    "item_code", "item", "group_name", "material_standard",
    "uom", "item_category", "specs",
]


def ensure_items(conn, records):
    """Upsert items from any source. `records` is an iterable of dicts holding
    at least 'item_code' plus any subset of ITEM_COLUMNS. Rows sharing an
    item_code are merged (first non-null value per attribute wins)."""
    merged = {}
    for rec in records:
        code = rec.get("item_code")
        if not code:
            continue
        cur = merged.setdefault(code, {c: None for c in ITEM_COLUMNS})
        cur["item_code"] = code
        for col in ITEM_COLUMNS:
            if col == "item_code":
                continue
            if cur[col] is None and rec.get(col) is not None:
                cur[col] = rec.get(col)
    rows = [tuple(m[c] for c in ITEM_COLUMNS) for m in merged.values()]
    if not rows:
        print("  items: nothing to ensure")
        return
    sql = (
        f"INSERT INTO items ({', '.join(ITEM_COLUMNS)}) VALUES %s "
        f"ON CONFLICT (item_code) DO NOTHING"
    )
    with conn.cursor() as cur:
        execute_values(cur, sql, rows, page_size=500)
    conn.commit()
    print(f"  items: ensured {len(rows)} distinct item codes (existing kept)")


def ensure_purchase_orders(conn, records):
    """Upsert purchase_order rows. `records` = dicts with 'po_number' and an
    optional 'po_date'."""
    merged = {}
    for rec in records:
        po = rec.get("po_number")
        if not po:
            continue
        cur = merged.setdefault(po, {"po_date": None})
        if cur["po_date"] is None and rec.get("po_date") is not None:
            cur["po_date"] = rec.get("po_date")
    rows = [(po, m["po_date"]) for po, m in merged.items()]
    if rows:
        with conn.cursor() as cur:
            execute_values(
                cur,
                "INSERT INTO purchase_order (po_number, po_date) "
                "VALUES %s ON CONFLICT (po_number) DO NOTHING",
                rows, page_size=500,
            )
        conn.commit()
    print(f"  purchase_order: ensured {len(rows)} PO numbers")


# ---------------------------------------------------------------------------
# FK maps (built from the DB after the parents are loaded)
# ---------------------------------------------------------------------------

def load_import_map(conn) -> dict:
    """Return {import_ref: import_id} so the import children resolve their FK."""
    with conn.cursor() as cur:
        cur.execute("SELECT import_ref, import_id FROM import_details")
        return {ref: i for ref, i in cur.fetchall() if ref is not None}


def load_shipment_map(conn) -> dict:
    """Return {(import_id, batch_no): shipment_id} to attach payments to a batch."""
    with conn.cursor() as cur:
        cur.execute("SELECT import_id, batch_no, shipment_id FROM shipment_details")
        return {(imp, bn): sid for imp, bn, sid in cur.fetchall()}


def load_shipment_by_import(conn) -> dict:
    """Return {import_id: shipment_id} (first shipment per import). Used to link
    a payment to its import's shipment when one import maps to one row/shipment."""
    with conn.cursor() as cur:
        cur.execute("SELECT import_id, shipment_id FROM shipment_details ORDER BY shipment_id")
        out = {}
        for imp, sid in cur.fetchall():
            out.setdefault(imp, sid)
        return out


# ---------------------------------------------------------------------------
# Currency resolution + live PKR conversion (imports)
# ---------------------------------------------------------------------------
# Currency symbol -> ISO. ¥ (U+00A5) is Chinese Yuan (CNY) in this data, not JPY.
_SYMBOL_TO_ISO = {"$": "USD", "¥": "CNY", "€": "EUR", "£": "GBP"}
# Blank-Currency rows carry the symbol in the FC cell's number format, e.g.
# [$$-409]=USD, [$¥-804]=CNY. pandas drops formats, so we read them via openpyxl.
_FC_FMT_CUR = re.compile(r"\[\$([^\-\]]+)-[0-9A-Fa-f]+\]")


def resolve_import_currencies(path=None) -> dict:
    """{df_row_index: ISO} for the real import rows. Uses the Currency-column
    symbol; when blank, reads the FC cell's [$sym-locale] format; defaults to USD."""
    path = path or IMPORT_FILE
    ws = openpyxl.load_workbook(path, data_only=True).active
    hdr = {ws.cell(row=IMPORT_HEADER_ROW + 1, column=c).value: c
           for c in range(1, ws.max_column + 1)}
    cur_c, fc_c = hdr.get("Currency"), hdr.get("Total Value(FC)")
    item_c, tab_c = hdr.get("Item Name"), hdr.get("Tab Status")
    out = {}
    for r in range(IMPORT_HEADER_ROW + 2, ws.max_row + 1):     # data starts Excel row 4
        if (ws.cell(row=r, column=item_c).value is None
                and ws.cell(row=r, column=tab_c).value is None):
            continue
        idx = r - (IMPORT_HEADER_ROW + 2)                       # -> DataFrame index
        sym = ws.cell(row=r, column=cur_c).value
        iso = None
        if isinstance(sym, str) and sym.strip():
            s = sym.strip()
            iso = _SYMBOL_TO_ISO.get(s) or _SYMBOL_TO_ISO.get(s[-1])
        else:                                                  # blank -> FC cell format
            m = _FC_FMT_CUR.search(ws.cell(row=r, column=fc_c).number_format or "")
            if m:
                iso = _SYMBOL_TO_ISO.get(m.group(1).strip()[-1])
            iso = iso or "USD"
        out[idx] = iso
    return out


_FALLBACK_PKR = {"USD": 278.0, "CNY": 41.0, "EUR": 300.0, "GBP": 350.0}


def get_pkr_rates(isos) -> dict:
    """{ISO: PKR per 1 unit} from today's live rates (open.er-api.com, keyless,
    includes PKR). Falls back to approximate hardcoded rates if the net is down."""
    isos = {c for c in isos if c}
    try:
        with urllib.request.urlopen(
                "https://open.er-api.com/v6/latest/USD", timeout=20) as resp:
            data = json.load(resp)
        rates = data["rates"]
        pkr_per_usd = rates["PKR"]
        out = {c: (1.0 if c == "PKR"
                   else pkr_per_usd / rates[c] if rates.get(c)
                   else _FALLBACK_PKR.get(c)) for c in isos}
        print("  forex (live " + data.get("time_last_update_utc", "")[:16] + "): "
              + ", ".join(f"{c}->PKR {out[c]:.3f}" for c in sorted(out) if out[c]))
        return out
    except Exception as exc:
        print(f"  !! forex live fetch failed ({type(exc).__name__}); using fallback")
        return {c: _FALLBACK_PKR.get(c) for c in isos}


# Re-export the cleaners so loaders can import everything from one place.
__all__ = [
    "IMPORT_FILE", "STOCK_FILE", "ISSUANCE_FILE", "STORE_REQ_FILE",
    "IMPORT_HEADER_ROW", "read_report", "read_import_rows", "import_ref_for",
    "ensure_items", "ensure_purchase_orders",
    "load_import_map", "load_shipment_map", "load_shipment_by_import",
    "resolve_import_currencies", "get_pkr_rates",
    "clean_text", "clean_number", "clean_int", "clean_date", "bulk_insert",
]
